"""
Reports Views — Filtered attendance reports with PDF and Excel export.

Reports are generated from live database queries — never cached.
Filters: week, service, semester.
"""
import io
import logging
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response

from apps.accounts.permissions import IsAdminOrAbove
from apps.students.models import Student
from apps.services.models import Semester
from apps.attendance.utils import calculate_attendance_percentage

logger = logging.getLogger(__name__)


class AttendanceReportView(APIView):
    """
    GET /api/reports/attendance/
    
    Filtered attendance report. Admin or above only.
    
    Query params:
    - semester_id: UUID (default: active semester)
    - service_group: S1/S2/S3 (optional filter)
    - week: ISO date string — filters to the week containing that date
    - below_threshold: true — only students below 70%
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        # Get semester
        semester_id = request.query_params.get('semester_id')
        if semester_id:
            try:
                semester = Semester.objects.get(id=semester_id)
            except Semester.DoesNotExist:
                return Response({'error': 'Semester not found.'}, status=404)
        else:
            semester = Semester.objects.filter(is_active=True).first()
            if not semester:
                return Response({'error': 'No active semester.'}, status=400)

        # Get students
        students_qs = Student.objects.filter(semester=semester)

        service_group = request.query_params.get('service_group')
        if service_group:
            students_qs = students_qs.filter(service_group=service_group)

        # Build report
        report = []
        below_count = 0

        for student in students_qs:
            pct_data = calculate_attendance_percentage(student, semester.id)
            item = {
                'student_id': str(student.id),
                'student_name': student.full_name,
                'matric_number': student.matric_number,
                'system_id': student.system_id,
                'service_group': student.service_group,
                **pct_data,
            }
            if pct_data['below_threshold']:
                below_count += 1

            report.append(item)

        # Filter below threshold only
        below_only = request.query_params.get('below_threshold')
        if below_only and below_only.lower() == 'true':
            report = [r for r in report if r['below_threshold']]

        # Sort by percentage ascending (worst first)
        report.sort(key=lambda r: r['percentage'])

        return Response({
            'semester_id': str(semester.id),
            'semester_name': semester.name,
            'total_students': len(report),
            'students_below_threshold': below_count,
            'report': report,
            'generated_at': timezone.now().isoformat(),
        })


class ExportPDFView(APIView):
    """
    GET /api/reports/export/pdf/
    
    Export attendance report as PDF.
    Uses ReportLab for PDF generation.
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return Response(
                {'error': 'PDF export library not available.'},
                status=500
            )

        # Get report data (reuse report logic)
        semester = Semester.objects.filter(is_active=True).first()
        if not semester:
            return Response({'error': 'No active semester.'}, status=400)

        students = Student.objects.filter(semester=semester)
        report_data = []
        for student in students:
            pct_data = calculate_attendance_percentage(student, semester.id)
            report_data.append({
                'name': student.full_name,
                'matric': student.matric_number or student.system_id,
                'group': student.service_group or '-',
                'valid': pct_data['valid_count'],
                'total': pct_data['total_required'],
                'pct': f"{pct_data['percentage']:.1f}%",
                'flag': '⚠️' if pct_data['below_threshold'] else '✓',
            })

        report_data.sort(key=lambda r: float(r['pct'].rstrip('%')))

        # Build PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(
            f'Attendance Report — {semester.name}',
            styles['Heading1']
        ))
        elements.append(Paragraph(
            f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))

        # Table
        header = ['Name', 'Matric/ID', 'Group', 'Valid', 'Total', '%', 'Status']
        data = [header] + [
            [r['name'], r['matric'], r['group'], r['valid'], r['total'], r['pct'], r['flag']]
            for r in report_data
        ]

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(table)

        doc.build(elements)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="attendance_report_{semester.name.replace(" ", "_")}.pdf"'
        )
        return response


class ExportExcelView(APIView):
    """
    GET /api/reports/export/excel/
    
    Export attendance report as Excel/CSV.
    Uses openpyxl for Excel generation.
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {'error': 'Excel export library not available.'},
                status=500
            )

        semester = Semester.objects.filter(is_active=True).first()
        if not semester:
            return Response({'error': 'No active semester.'}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'

        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        warning_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')

        # Title row
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Attendance Report — {semester.name}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}'

        # Headers
        headers = ['Name', 'Matric/ID', 'Group', 'Valid', 'Total', 'Percentage', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data
        students = Student.objects.filter(semester=semester)
        row = 5
        for student in students:
            pct_data = calculate_attendance_percentage(student, semester.id)
            ws.cell(row=row, column=1, value=student.full_name)
            ws.cell(row=row, column=2, value=student.matric_number or student.system_id)
            ws.cell(row=row, column=3, value=student.service_group or '-')
            ws.cell(row=row, column=4, value=pct_data['valid_count'])
            ws.cell(row=row, column=5, value=pct_data['total_required'])
            ws.cell(row=row, column=6, value=f"{pct_data['percentage']:.1f}%")
            ws.cell(row=row, column=7, value='BELOW 70%' if pct_data['below_threshold'] else 'OK')

            if pct_data['below_threshold']:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = warning_fill

            row += 1

        # Auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_length, 30)

        buffer = io.BytesIO()
        wb.save(buffer)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="attendance_report_{semester.name.replace(" ", "_")}.xlsx"'
        )
        return response


class DashboardStatsView(APIView):
    """
    GET /api/reports/dashboard-stats/

    Returns aggregated stats for admin dashboard charts:
    - Attendance by day (last 14 days)
    - Service group distribution
    - Sign-in time histogram (hourly buckets)
    - Weekly trend (last 8 weeks)
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        from collections import defaultdict
        from apps.attendance.models import AttendanceRecord
        from apps.services.models import Service

        semester = Semester.objects.filter(is_active=True).first()
        if not semester:
            return Response({'error': 'No active semester.'}, status=400)

        now = timezone.now()

        # ── Attendance by day (last 14 days) ──
        fourteen_ago = now - timedelta(days=14)
        daily_records = (
            AttendanceRecord.objects
            .filter(
                service__semester=semester,
                signed_in_at__gte=fourteen_ago,
                is_valid=True,
            )
            .values_list('signed_in_at', flat=True)
        )
        daily_counts = defaultdict(int)
        for ts in daily_records:
            daily_counts[ts.strftime('%Y-%m-%d')] += 1

        # Fill in missing days with 0
        attendance_by_day = []
        for i in range(14):
            day = (now - timedelta(days=13 - i)).strftime('%Y-%m-%d')
            attendance_by_day.append({'date': day, 'count': daily_counts.get(day, 0)})

        # ── Service group distribution ──
        group_distribution = {}
        for group_code in ['S1', 'S2', 'S3']:
            group_distribution[group_code] = Student.objects.filter(
                semester=semester, service_group=group_code, is_active=True
            ).count()

        # ── Sign-in time histogram (hourly buckets) ──
        all_signins = (
            AttendanceRecord.objects
            .filter(service__semester=semester, is_valid=True)
            .values_list('signed_in_at', flat=True)
        )
        hourly_counts = defaultdict(int)
        for ts in all_signins:
            hourly_counts[ts.hour] += 1
        signin_histogram = [
            {'hour': h, 'count': hourly_counts.get(h, 0)}
            for h in range(6, 22)  # 6 AM to 10 PM
        ]

        # ── Weekly trend (last 8 weeks) ──
        weekly_trend = []
        for w in range(7, -1, -1):
            week_start = now - timedelta(weeks=w + 1)
            week_end = now - timedelta(weeks=w)
            week_valid = AttendanceRecord.objects.filter(
                service__semester=semester,
                signed_in_at__gte=week_start,
                signed_in_at__lt=week_end,
                is_valid=True,
            ).count()
            week_total = AttendanceRecord.objects.filter(
                service__semester=semester,
                signed_in_at__gte=week_start,
                signed_in_at__lt=week_end,
            ).count()
            pct = (week_valid / week_total * 100) if week_total > 0 else 0
            weekly_trend.append({
                'week': week_end.strftime('%Y-%m-%d'),
                'valid': week_valid,
                'total': week_total,
                'percentage': round(pct, 1),
            })

        return Response({
            'attendance_by_day': attendance_by_day,
            'group_distribution': group_distribution,
            'signin_histogram': signin_histogram,
            'weekly_trend': weekly_trend,
        })


class SemesterComparisonView(APIView):
    """
    GET /api/reports/semester-comparison/

    Returns attendance summary for all semesters for cross-semester analytics.
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        from apps.services.models import Service
        from apps.attendance.models import AttendanceRecord
        from django.db.models import Q, Avg

        semesters = Semester.objects.all().order_by('start_date')
        comparison = []

        for sem in semesters:
            total_students = Student.objects.filter(semester=sem, is_active=True).count()
            total_services = Service.objects.filter(
                semester=sem, is_cancelled=False
            ).count()

            # Calculate avg attendance percentage across all students
            students = Student.objects.filter(semester=sem, is_active=True)
            if total_students > 0 and total_services > 0:
                total_valid = AttendanceRecord.objects.filter(
                    service__semester=sem, is_valid=True
                ).count()
                avg_pct = (total_valid / (total_students * total_services)) * 100 if total_students * total_services > 0 else 0
                below_count = 0
                for student in students:
                    pct = calculate_attendance_percentage(student, sem.id)
                    if pct['below_threshold']:
                        below_count += 1
            else:
                avg_pct = 0
                below_count = 0

            comparison.append({
                'semester_id': str(sem.id),
                'name': sem.name,
                'start_date': str(sem.start_date),
                'end_date': str(sem.end_date),
                'is_active': sem.is_active,
                'total_students': total_students,
                'total_services': total_services,
                'avg_percentage': round(avg_pct, 1),
                'below_threshold_count': below_count,
            })

        return Response({
            'semesters': comparison,
            'total_semesters': len(comparison),
        })


class StudentTrendView(APIView):
    """
    GET /api/reports/student-trend/?student_id=<uuid>

    Returns per-semester attendance data for a specific student.
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response({'error': 'student_id is required.'}, status=400)

        # Find all student records across semesters (same matric/phone)
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=404)

        # Find same student across semesters by matric or phone
        from django.db.models import Q
        match_q = Q(phone_number=student.phone_number)
        if student.matric_number:
            match_q |= Q(matric_number=student.matric_number)

        all_records = Student.objects.filter(match_q).select_related('semester').order_by('semester__start_date')

        trend = []
        for s in all_records:
            pct = calculate_attendance_percentage(s, s.semester_id)
            trend.append({
                'semester_id': str(s.semester_id),
                'semester_name': s.semester.name,
                'percentage': pct['percentage'],
                'valid_count': pct['valid_count'],
                'total_required': pct['total_required'],
                'below_threshold': pct['below_threshold'],
            })

        return Response({
            'student_name': student.full_name,
            'trend': trend,
        })


class ScanMetricsView(APIView):
    """
    GET /api/reports/scan-metrics/<service_id>/

    Computes scan speed metrics for a specific service:
    - Total scans
    - Avg time between scans per protocol member
    - Scans per 5-minute bucket over time
    - Per-protocol-member breakdown
    """
    permission_classes = [IsAdminOrAbove]

    def get(self, request, service_id):
        from collections import defaultdict
        from apps.attendance.models import AttendanceRecord
        from apps.services.models import Service

        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            return Response({'error': 'Service not found.'}, status=404)

        records = (
            AttendanceRecord.objects
            .filter(service=service)
            .select_related('protocol_member')
            .order_by('signed_in_at')
        )

        total_scans = records.count()
        if total_scans == 0:
            return Response({
                'service_id': str(service_id),
                'total_scans': 0,
                'avg_scans_per_minute': 0,
                'timeline': [],
                'per_member': [],
            })

        # ── Per-member breakdown ──
        member_records = defaultdict(list)
        for r in records:
            member_name = r.protocol_member.full_name if r.protocol_member else 'Unknown'
            member_records[member_name].append(r.signed_in_at)

        per_member = []
        for name, timestamps in member_records.items():
            timestamps.sort()
            gaps = []
            for i in range(1, len(timestamps)):
                gap = (timestamps[i] - timestamps[i - 1]).total_seconds()
                gaps.append(gap)
            avg_gap = sum(gaps) / len(gaps) if gaps else 0
            per_member.append({
                'name': name,
                'scan_count': len(timestamps),
                'avg_gap_seconds': round(avg_gap, 1),
            })
        per_member.sort(key=lambda x: x['scan_count'], reverse=True)

        # ── Timeline (5-min buckets) ──
        all_timestamps = list(records.values_list('signed_in_at', flat=True))
        if all_timestamps:
            first = min(all_timestamps)
            last = max(all_timestamps)
            total_minutes = max((last - first).total_seconds() / 60, 1)
            avg_scans_per_minute = round(total_scans / total_minutes, 2)

            # Build 5-min buckets
            timeline = []
            bucket_counts = defaultdict(int)
            for ts in all_timestamps:
                # Round down to 5-min bucket
                bucket = ts.replace(
                    minute=(ts.minute // 5) * 5,
                    second=0,
                    microsecond=0,
                )
                bucket_counts[bucket.isoformat()] += 1

            for bucket_key in sorted(bucket_counts.keys()):
                timeline.append({
                    'time': bucket_key,
                    'count': bucket_counts[bucket_key],
                })
        else:
            avg_scans_per_minute = 0
            timeline = []

        return Response({
            'service_id': str(service_id),
            'total_scans': total_scans,
            'avg_scans_per_minute': avg_scans_per_minute,
            'timeline': timeline,
            'per_member': per_member,
        })

